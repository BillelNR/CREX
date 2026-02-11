import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from concurrent.futures import ThreadPoolExecutor
import time

st.set_page_config(
    page_title="Transavia - Traitement CREX",
    page_icon="✈️",
    layout="wide"
)

# CSS personnalisé - Thème TransaviaFR
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .main-header {
        font-size: 2.2rem;
        color: #003366;
        text-align: center;
        margin-bottom: 0.5rem;
        font-weight: 700;
        letter-spacing: -0.5px;
    }
    
    .sub-header {
        font-size: 1rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: 400;
    }
    
    .upload-section {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        padding: 2rem;
        border-radius: 10px;
        border: 2px dashed #003366;
        margin: 1rem 0;
    }
    
    .file-info-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #003366;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin: 0.5rem 0;
    }
    
    .success-card {
        background: linear-gradient(135deg, #e6f7e9 0%, #d4edda 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #28a745;
        margin: 1rem 0;
    }
    
    .stButton > button {
        background: linear-gradient(135deg, #003366 0%, #00509e 100%);
        color: white;
        font-weight: 600;
        border: none;
        padding: 0.8rem 2rem;
        border-radius: 8px;
        font-size: 1rem;
        transition: all 0.3s ease;
        width: 100%;
    }
    
    .stButton > button:hover {
        background: linear-gradient(135deg, #002244 0%, #003366 100%);
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,51,102,0.2);
    }
    
    .download-btn {
        background: linear-gradient(135deg, #FF6600 0%, #ff8533 100%) !important;
    }
    
    .download-btn:hover {
        background: linear-gradient(135deg, #e65c00 0%, #ff6600 100%) !important;
    }
    
    .sidebar-title {
        color: #003366;
        font-weight: 700;
        font-size: 1.2rem;
        margin-bottom: 1rem;
    }
    
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        text-align: center;
    }
    
    .metric-value {
        font-size: 1.8rem;
        font-weight: 700;
        color: #003366;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: #666;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    .logo-container {
        text-align: center;
        margin-bottom: 2rem;
    }
    
    .info-box {
        background: #e6f2ff;
        padding: 1rem;
        border-radius: 8px;
        border-left: 3px solid #003366;
        margin: 1rem 0;
    }
    
    .password-box {
        background: #fff3cd;
        padding: 1rem;
        border-radius: 8px;
        border-left: 3px solid #ffc107;
        margin-top: 2rem;
    }
    
    .password-label {
        color: #856404;
        font-size: 0.9rem;
        font-weight: 600;
        margin-bottom: 0.5rem;
    }
    
    .password-value {
        background: white;
        padding: 0.75rem;
        border-radius: 5px;
        border: 1px solid #ffc107;
        font-family: 'Courier New', monospace;
        font-weight: bold;
        color: #003366;
        text-align: center;
        letter-spacing: 1px;
    }
    
    .filename-info {
        background: #e6f7ff;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #003366;
        margin: 1rem 0;
    }
    
    .progress-bar {
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def format_date_french(date_obj):
    """Format date en français de manière optimisée"""
    if pd.isna(date_obj):
        return ""
    try:
        if isinstance(date_obj, pd.Timestamp):
            days_fr = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
            months_fr = ["janvier", "février", "mars", "avril", "mai", "juin", 
                        "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
            return f"{days_fr[date_obj.weekday()]} {date_obj.day} {months_fr[date_obj.month - 1]}"
    except:
        pass
    return str(date_obj)

def traiter_feuille_optimise(sheet_name, df):
    """Traitement optimisé d'une feuille Excel"""
    if len(df) < 2:
        return []
    
    # Filtrer les lignes qui ne sont pas des titres
    mask = pd.Series([True] * len(df), index=df.index)
    
    # Analyser la colonne 0 pour les titres
    if 0 in df.columns:
        col0 = df.iloc[:, 0].astype(str).str.strip().str.upper()
        
        # Titres à exclure
        titres_exclus = {"AUTRES", "PRODUIT", "SERVICE", "EQUIPEMENT", "AUTRE"}
        mask_titre = col0.isin(titres_exclus) | (col0.str.len() > 50)
        
        # Exclure les titres et la première ligne (en-tête)
        mask = ~mask_titre
        mask.iloc[0] = False  # Exclure l'en-tête
    
    # Appliquer le masque
    df_filtered = df[mask].copy()
    
    if len(df_filtered) == 0:
        return []
    
    # Convertir les dates de manière vectorisée
    dates_valides = []
    for idx, row in df_filtered.iterrows():
        try:
            date_val = pd.to_datetime(row.iloc[0] if 0 in row.index else None)
            if pd.notna(date_val):
                dates_valides.append(True)
            else:
                dates_valides.append(False)
        except:
            dates_valides.append(False)
    
    df_filtered = df_filtered[dates_valides].copy()
    
    if len(df_filtered) == 0:
        return []
    
    # Extraire les données de manière optimisée
    data_rows = []
    for idx, row in df_filtered.iterrows():
        try:
            date_val = pd.to_datetime(row.iloc[0])
            data_row = {
                'Date Vol': date_val,
                'Aircraft Registration': row.iloc[2] if len(row) > 2 else '',
                'Flight Number': row.iloc[3] if len(row) > 3 else '',
                'Origin': row.iloc[4] if len(row) > 4 else '',
                'Destination': row.iloc[5] if len(row) > 5 else '',
                'Catering': row.iloc[7] if len(row) > 7 else '',
                'Non Conformité': row.iloc[8] if len(row) > 8 else '',
                'Event Title': row.iloc[9] if len(row) > 9 else '',
                'General Remarks': row.iloc[10] if len(row) > 10 else '',
            }
            data_rows.append(data_row)
        except:
            continue
    
    return data_rows

def traiter_exactement_comme_vba(uploaded_file, progress_bar=None):
    """Version optimisée du traitement VBA"""
    try:
        start_time = time.time()
        
        # Lire toutes les feuilles en une fois
        xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
        
        # Filtrer les feuilles (sauf EXPORT)
        sheet_names = [name for name in xls.sheet_names if name.upper() != "EXPORT"]
        
        all_data = []
        
        if progress_bar:
            progress_bar.progress(10, text="Lecture des feuilles...")
        
        # Traiter chaque feuille
        for i, sheet_name in enumerate(sheet_names):
            try:
                # Lire la feuille sans en-tête pour traiter toutes les lignes
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine='openpyxl')
                
                # Traiter la feuille
                data_rows = traiter_feuille_optimise(sheet_name, df)
                all_data.extend(data_rows)
                
                if progress_bar and sheet_names:
                    progress_value = 10 + int((i + 1) / len(sheet_names) * 40)
                    progress_bar.progress(progress_value, text=f"Traitement feuille {i+1}/{len(sheet_names)}...")
                    
            except Exception as e:
                st.warning(f"Erreur sur la feuille {sheet_name}: {str(e)}")
                continue
        
        if not all_data:
            return None, "Aucune donnée valide trouvée dans le fichier.", None
        
        if progress_bar:
            progress_bar.progress(60, text="Organisation des données...")
        
        # Créer DataFrame
        df_all = pd.DataFrame(all_data)
        
        # Nettoyer les données
        origins_valides = {"ORY", "MRS", "LYS", "NTE", "BRU", "MPL", "RNS", "BOD", "TLS"}
        
        # Optimiser la création des feuilles par origine
        nouvelles_feuilles = {}
        
        # Grouper par origine de manière vectorisée
        if 'Origin' in df_all.columns:
            df_all['Origin_clean'] = df_all['Origin'].astype(str).str.strip().str.upper()
            
            # Séparer les origines valides et "Autre"
            mask_valide = df_all['Origin_clean'].isin(origins_valides)
            df_all['Sheet_Name'] = df_all['Origin_clean'].where(mask_valide, "Autre")
            
            # Grouper par nom de feuille
            grouped = df_all.groupby('Sheet_Name')
            
            for sheet_name, group in grouped:
                nouvelles_feuilles[sheet_name] = []
                
                # Préparer les données pour cette feuille
                for _, row in group.iterrows():
                    ligne_complete = {
                        'Date Vol': row['Date Vol'],
                        'Aircraft Registration': row['Aircraft Registration'],
                        'Flight Number': row['Flight Number'],
                        'Origin': row['Origin'],
                        'Destination': row['Destination'],
                        'Catering': row['Catering'],
                        'Non Conformité': row['Non Conformité'],
                        'Event Title': row['Event Title'],
                        'General Remarks': row['General Remarks'],
                        'Accepté/Refusé': None,
                        'Commentaire': None,
                        'Autre': None,
                        'KAM / TO': None,
                        'Commentaire_2': None
                    }
                    nouvelles_feuilles[sheet_name].append(ligne_complete)
        else:
            # Fallback si pas de colonne Origin
            nouvelles_feuilles["Autre"] = []
            for _, row in df_all.iterrows():
                ligne_complete = {
                    'Date Vol': row['Date Vol'],
                    'Aircraft Registration': row['Aircraft Registration'],
                    'Flight Number': row['Flight Number'],
                    'Origin': row.get('Origin', ''),
                    'Destination': row.get('Destination', ''),
                    'Catering': row['Catering'],
                    'Non Conformité': row['Non Conformité'],
                    'Event Title': row['Event Title'],
                    'General Remarks': row['General Remarks'],
                    'Accepté/Refusé': None,
                    'Commentaire': None,
                    'Autre': None,
                    'KAM / TO': None,
                    'Commentaire_2': None
                }
                nouvelles_feuilles["Autre"].append(ligne_complete)
        
        if progress_bar:
            progress_bar.progress(80, text="Création de la consolidation...")
        
        # Créer la feuille Consolidation avec formules
        consolidation_data = []
        for sheet_name, data in nouvelles_feuilles.items():
            for idx in range(len(data)):
                row_num = idx + 2
                row_copy = {
                    'Date Vol': f"='{sheet_name}'!A{row_num}",
                    'Aircraft Registration': f"='{sheet_name}'!B{row_num}",
                    'Flight Number': f"='{sheet_name}'!C{row_num}",
                    'Origin': f"='{sheet_name}'!D{row_num}",
                    'Destination': f"='{sheet_name}'!E{row_num}",
                    'Catering': f"='{sheet_name}'!F{row_num}",
                    'Non Conformité': f"='{sheet_name}'!G{row_num}",
                    'Event Title': f"='{sheet_name}'!H{row_num}",
                    'General Remarks': f"='{sheet_name}'!I{row_num}",
                    'Accepté/Refusé': f"='{sheet_name}'!J{row_num}",
                    'Commentaire': f"='{sheet_name}'!K{row_num}",
                    'Autre': f"='{sheet_name}'!L{row_num}",
                    'KAM / TO': f"='{sheet_name}'!M{row_num}",
                    'Commentaire_2': f"='{sheet_name}'!N{row_num}",
                }
                consolidation_data.append(row_copy)
        
        nouvelles_feuilles['Consolidation'] = consolidation_data
        
        if progress_bar:
            progress_bar.progress(90, text="Génération du fichier Excel...")
        
        # Créer le fichier Excel
        excel_output = creer_excel_avec_formatage_optimise(nouvelles_feuilles)
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        if progress_bar:
            progress_bar.progress(100, text=f"Terminé en {processing_time:.1f} secondes")
        
        return excel_output, None, df_all
        
    except Exception as e:
        return None, f"Erreur lors du traitement: {str(e)}", None

def creer_excel_avec_formatage_optimise(nouvelles_feuilles):
    """Version optimisée de la création Excel + protection (Option A) avec colonnes J et K déverrouillées"""
    try:
        output = BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        
        # Préparer les styles une seule fois
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        headers = [
            "Date Vol", "Aircraft Registration", "Flight Number", "Origin", "Destination",
            "Catering", "Non Conformité", "Event Title", "General Remarks",
            "Accepté/Refusé", "Commentaire", "Autre", "KAM / TO", "Commentaire_2"
        ]
        
        for sheet_name, data in nouvelles_feuilles.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Ajouter les en-têtes
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True, color="003366")
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.fill = openpyxl.styles.PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            # Ajouter les données
            if isinstance(data, list):
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, col_name in enumerate(headers, 1):
                        value = row_data.get(col_name)

                        # Si c'est une formule Excel (ex: "='Feuille'!A2"), on l'écrit telle quelle
                        if isinstance(value, str) and value.startswith("="):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            # Formater les dates
                            if col_name == "Date Vol" and value is not None:
                                try:
                                    # Ne formater que si c'est une vraie date, pas une formule
                                    if isinstance(value, pd.Timestamp):
                                        value = format_date_french(value)
                                    else:
                                        # tenter conversion si type date-like
                                        value = format_date_french(pd.to_datetime(value))
                                except:
                                    pass
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)

                        cell.border = thin_border
                        
                        # Appliquer les alignements
                        if col_idx in [6, 7, 8, 9, 10, 11, 12, 13, 14]:
                            cell.alignment = wrap_alignment
                        elif col_idx == 1:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        else:
                            cell.alignment = Alignment(vertical='top')
            
            # Ajuster les largeurs de colonnes
            column_widths = {
                'A': 20, 'B': 15, 'C': 15, 'D': 10, 'E': 10,
                'F': 30, 'G': 30, 'H': 30, 'I': 50,
                'J': 15, 'K': 30, 'L': 20, 'M': 15, 'N': 30
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Ajouter les validations de données si nécessaire
            if isinstance(data, list) and len(data) > 0:
                last_row = len(data) + 1
                if last_row >= 2:
                    # Validation pour la colonne J
                    dv_j = DataValidation(
                        type="list",
                        formula1='"Accepté,Refusé,N/A"',
                        allow_blank=True
                    )
                    dv_j.add(f'J2:J{last_row}')
                    ws.add_data_validation(dv_j)
                    
                    # Validation pour la colonne M
                    dv_m = DataValidation(
                        type="list",
                        formula1='"Accepté,Refusé,N/A"',
                        allow_blank=True
                    )
                    dv_m.add(f'M2:M{last_row}')
                    ws.add_data_validation(dv_m)

                    # 🔓 Déverrouiller colonnes J et K (10 et 11) pour toutes les lignes de données
                    for col in [10, 11]:
                        for r in range(2, last_row + 1):
                            ws.cell(row=r, column=col).protection = Protection(locked=False)
            
            # 🔒 Protection de la feuille (Option A)
            ws.protection.sheet = True
            # facultatif selon versions : ws.protection.enable()
            ws.protection.set_password('newrest2025')
        
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"Erreur lors de la création du fichier: {str(e)}")
        return None

def main():
    # Header avec logo Transavia
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div class="logo-container">', unsafe_allow_html=True)
        st.markdown('<h1 class="main-header">✈️ TRANSFORMA CREX</h1>', unsafe_allow_html=True)
        st.markdown('<p class="sub-header">Outil de traitement automatisé des rapports CREX</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown('<p class="sidebar-title">📋 Informations</p>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="info-box">
        <strong>Fonctionnalités :</strong>
        <ul style="margin-top: 10px; padding-left: 20px;">
            <li>Traitement automatique des données</li>
            <li>Création de feuilles par origine</li>
            <li>Formatage des dates en français</li>
            <li>Ajout de listes déroulantes</li>
            <li>Consolidation automatique</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
        
        # Section mot de passe ajoutée ici
        st.markdown('<div class="password-box">', unsafe_allow_html=True)
        st.markdown('<div class="password-label">🔐 Mot de passe du fichier :</div>', unsafe_allow_html=True)
        st.markdown('<div class="password-value">newrest2025</div>', unsafe_allow_html=True)
        st.markdown('<p style="font-size: 0.8rem; color: #856404; margin-top: 0.5rem;">À utiliser pour ouvrir le fichier généré</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Informations de performance
        st.markdown("---")
        st.markdown("""
        <div style="margin-top: 1rem;">
            <p style="font-size: 0.9rem; color: #666;"><strong>⚡ Performances :</strong></p>
            <p style="font-size: 0.8rem; color: #666;">• Traitement optimisé</p>
            <p style="font-size: 0.8rem; color: #666;">• Lecture vectorisée</p>
            <p style="font-size: 0.8rem; color: #666;">• Pas de délais</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        st.markdown("""
        <div style="text-align: center; margin-top: 2rem;">
            <p style="color: #666; font-size: 0.8rem;">Version 1.1 (Optimisée)</p>
            <p style="color: #666; font-size: 0.8rem;">© 2024 Transavia France</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Zone d'upload principale
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    
    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown("### 📤 Déposer votre fichier CREX")
        st.markdown("Téléversez le fichier Excel à transformer")
    with col2:
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.markdown('<div class="metric-value">.xlsx</div>', unsafe_allow_html=True)
        st.markdown('<div class="metric-label">Format accepté</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "",
        type=['xlsx'],
        label_visibility="collapsed",
        help="Sélectionnez votre fichier Excel à traiter"
    )
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_file is not None:
        # Informations sur le fichier
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown('<div class="file-info-card">', unsafe_allow_html=True)
            st.markdown(f"**📄 Fichier source :** {uploaded_file.name}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown('<div class="file-info-card">', unsafe_allow_html=True)
            file_size_mb = uploaded_file.size / (1024 * 1024)
            st.markdown(f"**💾 Taille :** {file_size_mb:.2f} MB")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col3:
            st.markdown('<div class="file-info-card">', unsafe_allow_html=True)
            st.markdown("**📊 Statut :** Prêt")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Bouton de traitement avec indicateur de progression
        if st.button("🚀 Lancer le traitement (Version rapide)", type="primary"):
            progress_bar = st.progress(0, text="Initialisation...")
            
            with st.spinner("Traitement optimisé en cours..."):
                excel_output, erreur, df_data = traiter_exactement_comme_vba(uploaded_file, progress_bar)
            
            progress_bar.empty()
            
            if erreur:
                st.error(f"⚠️ {erreur}")
            elif excel_output:
                st.markdown('<div class="success-card">', unsafe_allow_html=True)
                
                col_s1, col_s2, col_s3 = st.columns(3)
                
                with col_s1:
                    st.metric("✅ Traitement terminé", "Succès")
                
                with col_s2:
                    total_lignes = len(df_data) if df_data is not None else 0
                    st.metric("📈 Lignes traitées", f"{total_lignes:,}")
                
                with col_s3:
                    if df_data is not None and 'Origin' in df_data.columns:
                        feuilles = len(set(df_data['Origin'].dropna().astype(str).str.strip().str.upper())) + 2
                    else:
                        feuilles = 1
                    st.metric("📑 Feuilles créées", feuilles)
                
                st.markdown('</div>', unsafe_allow_html=True)
                
                # Téléchargement avec le même nom que le fichier d'entrée
                st.markdown("### 📥 Télécharger le résultat")
                
                # Préparation du nom de fichier (même nom que l'entrée)
                input_filename = uploaded_file.name
                # Assurer l'extension .xlsx
                if not input_filename.lower().endswith('.xlsx'):
                    input_filename = f"{input_filename}.xlsx"
                
                # Afficher l'info sur le nom du fichier
                st.markdown(f"""
                <div class="filename-info">
                <strong>📝 Nom du fichier de sortie :</strong><br>
                <code>{input_filename}</code>
                </div>
                """, unsafe_allow_html=True)
                
                col_d1, col_d2 = st.columns([3, 1])
                with col_d1:
                    st.info(f"Le fichier traité '{input_filename}' est prêt à être téléchargé")
                
                with col_d2:
                    st.download_button(
                        label="📥 Télécharger",
                        data=excel_output,
                        file_name=input_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="secondary"
                    )
    else:
        # Instructions quand aucun fichier n'est uploadé
        st.markdown("""
        <div style="text-align: center; padding: 3rem; color: #666;">
            <h3 style="color: #003366;">📋 Instructions</h3>
            <p>1. Préparez votre fichier Excel CREX</p>
            <p>2. Cliquez sur "Browse files" ou glissez-déposez</p>
            <p>3. Lancez le traitement automatique</p>
            <p>4. Téléchargez le résultat (même nom que l'original)</p>
            <br>
            <p style="color: #003366; font-weight: 600;">⚡ Version optimisée pour les performances</p>
        </div>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()